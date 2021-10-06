#! /bin/python

import openpyxl
import sys

args = sys.argv

# エクセルファイルの読み込み
wb = openpyxl.load_workbook("./test.xlsx")
# ワークシートを取得
ws = wb['Sheet1']

# 最終列を取得
max_row = ws.max_row

# 確認する data = [args[1], args[2], args[3], args[4]]
c1 = ws.cell(max_row, 1)
c2 = ws.cell(max_row, 2)
c3 = ws.cell(max_row, 3)
c4 = ws.cell(max_row, 4)

day = args[1]
if1 = "c1.value == day"
if2 = "c2.value == int(args[2])"
if3 = "c3.value == int(args[3])"
if4 = "c4.value == int(args[4])"

if if1 and if2 and if3 and if4:
    print("エクセルへの書き込みを確認しました")
else:
    print("エクセルの値に間違いがあります")
    print("確認してください")

# 確認したエクセルファイルを閉じる
wb.close()
