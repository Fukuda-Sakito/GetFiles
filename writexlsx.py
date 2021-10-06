#! /usr/bin/env python

import openpyxl
import sys
import shutil
import os
args = sys.argv

# エクセルファイルの読み込み
wb = openpyxl.load_workbook("./GetFiles統計.xlsx")
# ワークシートを取得
ws = wb['Sheet1']

# 編集前にバックアップを取る．バックアップは5回分作成
i = 4
j = 1
if os.path.exists("./backup_xlsx/GetFiles統計(5).xlsx"):
    os.remove("./backup_xlsx/GetFiles統計(1).xlsx")
    while j < 5:
        os.rename(f"./backup_xlsx/GetFiles統計({j + 1}).xlsx", f"./backup_xlsx/GetFiles統計({j}).xlsx")
        j += 1
    shutil.copy("./GetFiles統計.xlsx", "./backup_xlsx/GetFiles統計(5).xlsx")

# 最終列を取得
max_row = ws.max_row

# 書き込む data = [args[1], args[2], args[3], args[4]]
c1 = ws.cell(max_row + 1, 1)
c2 = ws.cell(max_row + 1, 2)
c3 = ws.cell(max_row + 1, 3)
c4 = ws.cell(max_row + 1, 4)

day = args[1]
c1.value = day
c2.value = int(args[2])
c3.value = int(args[3])
c4.value = int(args[4])

# 上書き保存
wb.save("./GetFiles統計.xlsx")
